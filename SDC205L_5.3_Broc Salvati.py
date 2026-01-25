# SDC205 Spreadsheet Automation Project (Generate Report Extension)
# Student ID: Brosal3735

import csv
from datetime import datetime

import openpyxl
from openpyxl.chart import BarChart, LineChart, Reference


STUDENT_ID = "Brosal3735"


def _read_csv_rows(csv_path):
    """
    Reads CSV file with columns: Date, Original, Converted

    Returns:
        dates (list[str])
        initial_vals (list[float])
        converted_vals (list[float])
    """

    dates = []
    initial_vals = []
    converted_vals = []

    with open(csv_path, "r", newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)

        # If no header, fall back to regular reader
        if reader.fieldnames is None:
            f.seek(0)
            raw = csv.reader(f)

            for row in raw:
                if not row or len(row) < 3:
                    continue

                try:
                    dates.append(str(row[0]).strip())
                    initial_vals.append(float(row[1]))
                    converted_vals.append(float(row[2]))
                except ValueError:
                    continue

            return dates, initial_vals, converted_vals

        # Normal header case
        for row in reader:
            if not row:
                continue

            date_str = str(row.get("Date", "")).strip()

            if not date_str:
                continue

            try:
                init_val = float(row.get("Original", ""))
                conv_val = float(row.get("Converted", ""))
            except ValueError:
                continue

            dates.append(date_str)
            initial_vals.append(init_val)
            converted_vals.append(conv_val)

    return dates, initial_vals, converted_vals


def _write_excel_with_chart(dates, values, y_label, chart_type):
    """
    Writes data to Excel and creates chart
    """

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"

    # Headers
    ws["A1"] = "Date"
    ws["B1"] = y_label

    # Data
    for i in range(len(dates)):
        ws.cell(row=i + 2, column=1, value=dates[i])
        ws.cell(row=i + 2, column=2, value=values[i])

    max_row = len(dates) + 1

    labels = Reference(ws, min_col=1, min_row=2, max_row=max_row)
    data = Reference(ws, min_col=2, min_row=1, max_row=max_row)

    # Chart type
    if chart_type.lower() == "bar":
        chart = BarChart()
    else:
        chart = LineChart()

    # Title
    today = datetime.now().strftime("%m/%d/%Y")
    chart.title = f"{STUDENT_ID} {today}"

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)

    # Axis labels
    chart.x_axis.title = "Date"
    chart.y_axis.title = y_label

    # Add chart
    ws.add_chart(chart, "D2")

    wb.save("final.xlsx")


# createChart(csv_path: str, chart_type: str) -> None
# Required arguments:
#   csv_path (str): Path to the CSV data file
#   chart_type (str): "line" or "bar"
# Return value:
#   None
def createChart(csv_path, chart_type):

    dates, initial_vals, converted_vals = _read_csv_rows(csv_path)

    if not dates:
        print("No valid data found in the CSV file.")
        return

    print("\nChoose the data source:")
    print("1) Initial data (pounds/inches/Fahrenheit)")
    print("2) Converted data (kilograms/centimeters/Celsius)")

    choice = input("Enter 1 or 2: ").strip()

    if choice == "2":
        values = converted_vals
        y_label = "Converted Value"
    else:
        values = initial_vals
        y_label = "Initial Value"

    _write_excel_with_chart(dates, values, y_label, chart_type)

    print("\nReport generated: final.xlsx\n")


# generateReport(csv_path: str) -> None
# Required arguments:
#   csv_path (str): Path to the CSV data file
# Return value:
#   None
def generateReport(csv_path):

    print("\nGenerate Report")
    print("1) Line chart")
    print("2) Bar chart")

    chart_choice = input("Choose chart type (1 or 2): ").strip()

    if chart_choice == "2":
        chart_type = "bar"
    else:
        chart_type = "line"

    createChart(csv_path, chart_type)


def inputData():
    print("\nInput Data: (Use your previous logic here)\n")


def readData():
    print("\nRead Data: (Use your previous logic here)\n")


def menu():
    print("1) Input Data")
    print("2) Read Data")
    print("3) Generate Report")
    print("4) Exit")


def main():

    # CSV must be in same folder as this file
    csv_path = "project_data.csv"

    while True:

        menu()

        choice = input("Enter your choice: ").strip()

        if choice == "1":
            inputData()

        elif choice == "2":
            readData()

        elif choice == "3":
            generateReport(csv_path)

        elif choice == "4":
            print("Goodbye.")
            break

        else:
            print("Invalid option. Try again.\n")


if __name__ == "__main__":
    main()


